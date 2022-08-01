# This is parser. It scraps prices and other data from exist.ua for given list
# of articles in excel file (in the script folder) and stores data into new
# result excel file.
# This parsing has simple approach to read web-server's response - without
# js-content analysis and without VPN hiding.
# Specifics of parsing from exist.ua - this site responds randomly with 2
# different formats on 2 pages and needed data is in 2 divs.
# Excel file with requested list of articles must be placed in the same folder
# with parser file, and it is an export destination for respond excel file.

import logging  # library enables creation of logs
import time  # library enables work with data in time format
import datetime  # library enables work with data in date format
from random import randint  # library enables creation of random figures

import openpyxl  # library enables MS Excel files reading
import requests  # library enables getting a response from web-server
import xlwt  # library enables creation of MS Excel 2003 files
from bs4 import BeautifulSoup as BS  # library enables web pages parsing

# imitate web-session and query to browser and activate coockies
session = requests.Session()
# create tracing the logs
logger = logging.getLogger('exist_parser')
logger.setLevel(logging.DEBUG)

# imitate queries form different browsers not to let website ban the parser
headers = [
    {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.9; rv:45.0) Gecko/20100101 Firefox/45.0'},
    {'User-Agent': 'Mozilla/5.0 (Windows NT 5.1; rv:47.0) Gecko/20100101 Firefox/47.0','Accept':'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8'},
    {'User-Agent': 'Mozilla/5.0 (Windows NT 5.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/49.0.2623.112 Safari/537.36','Accept':'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8'},
    {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; rv:53.0) Gecko/20100101 Firefox/53.0','Accept':'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8'}
    ]

def parse_div(div):
    '''This function parses <div> which exists on both from two random pages in
    site's respond and contains all needed data.

    Attributes
    ----------
        - div: a unique div which contains data we need
    '''
    # knowing that needed data stored in <div> with class "rowOffers" (after
    # manual website analysis), save all such divs into the BS-object "rows"
    rows = div.find_all('div', attrs={'class': 'rowOffers'})
    # go through all sub-divs with class "rowOffers" and extract needed content
    for row in rows:
        # store 1-st part of needed content from corresponding div (with brand,
        # part number and description) to BS-object for further parsing
        brand_part_descr_div = row.find(
            'div', attrs={'class': 'row--search-result-name'})
        # get brand name as text by corresponding class,
        # and delete some trash like spaces and tags
        brand_bid = brand_part_descr_div.find(
            'div', attrs={'class': 'art'}).text.replace(
                '\n', ' ').strip()
        # get part number as text by corresponding class,
        # and delete some trash like spaces and tags
        part_number_bid = brand_part_descr_div.find(
            'div', attrs={'class': 'partno'}).text.replace(
                '\n', ' ').strip()
        # get description as text by corresponding class,
        # and delete some trash like spaces and tags
        description = brand_part_descr_div.find(
            'div', attrs={'class': 'descr'}).text.replace(
                '\n', ' ').strip()
        # store second part of needed content from corresponding <div> (with
        # quantity, availability date and price) to BS-object for further parse
        stock_date_price_div = row.find_all(
            'div', attrs={'class': 'pricerow'})
        # create temporary dictionary to store values from second div and to
        # further append to main dictionary
        temporary = []
        # check the data available in second div
        if stock_date_price_div:
            # go through sub-divs and extract needed content
            for offer in stock_date_price_div:
                # get stock quantity as text by corresponding class,
                # and delete some trash like spaces and tags
                stock = offer.find(
                    'span', attrs={'class': 'avail'}).text.replace(
                        '\n', ' ').strip()
                # get available date as text by corresponding class,
                # and delete some trash like spaces and tags
                date = offer.find(
                    'div', attrs={'class': 'stock-info'}).p.text.replace(
                        '\n', ' ').strip()
                # get price as text by corresponding class,
                # and delete some trash like spaces and tags
                price = offer.find(
                    'span', attrs={'class': 'price'}).text.replace(
                        'грн', ' ').replace(' ','').strip()
                # store parsed data from second div to temporary dictionary
                temporary.append(
                    {'stock': stock, 'date': date, 'price': price}
                )
        # store parsed data to main dictionary
        data.append(
            {
                'part_number_ask': part_number_ask,
                'part_number_bid': part_number_bid,
                'brand_bid': brand_bid,
                'description': description,
                'stock_date_price_div': temporary
            }
        )
# =========================== def parse_div end ================================

def write_to_excel(data, file_name='response_exist.xls'):
    '''Create table with parsed data

    Attributes
    ----------
        - data: a list of dictionaries which will be saved to an excel file
        - file_name
    '''
    book = xlwt.Workbook(encoding='utf-8')  # create MS Excel file
    sheet = book.add_sheet('EXIST')  # create sheet and name it
    row_num = 0  # indicate the first row to write in the sheet
    font_style = xlwt.XFStyle()  # set style for text
    font_style.font.bold = True  # set text in bold
    row = sheet.row(row_num)  # indicate number of row to work with

    # create table headings ("column #", "content", "text style")
    row.write(0,'date', font_style)
    row.write(1,'source', font_style)
    row.write(2,'brand_ask', font_style)
    row.write(3,'part_number_ask', font_style)
    row.write(4,'part_number_bid', font_style)
    row.write(5,'brand_bid', font_style)
    row.write(6,'quality', font_style)
    row.write(7,'description', font_style)
    row.write(8,'availability', font_style)
    row.write(9,'price_UAH', font_style)
    font_style = xlwt.XFStyle()
    row_num +=1
    today = datetime.datetime.today().strftime('%d.%m.%Y')

    # filling up the table with parsed data row by row in a loop
    for d in data:
        for price in d['stock_date_price_div']:
            row = sheet.row(row_num)
            row.write(0,today)
            row.write(1,domain[8:])
            row.write(2,brand_ask)
            row.write(3,d['part_number_ask'])
            row.write(4,d['part_number_bid'])
            row.write(5,d['brand_bid'])
            if brand_ask.upper() == d['brand_bid'].upper():
                row.write(6,'genuine')
            elif d['part_number_bid'] == 'N/A':
                row.write(6,'N/A')
            else:
                row.write(6,'aftermarket')
            row.write(7,d['description'])
            row.write(8,price['date'])
            row.write(9,float(price['price']))
            row_num +=1
    book.save(file_name) #  save excel file

# ========================= def write_to_excel end =============================

# indicate requested brand name
brand_ask = 'Mitsubishi'
# set first part of the link to get address for parser
domain = 'https://exist.ua'
# create empty dictionary to store found data
data = []
# create empty temporary dictionary to store found data from second
# corresponding <div>
temporary = []
# open excel file with a list of requested part numbers
book = openpyxl.load_workbook('request.xlsx')
# activate excel sheet to work with
sheet = book.active
for row in sheet.iter_rows(min_row=0, min_col=0, max_col=2):
    part_number_ask = str(row[0].value).strip()  # read part number from Excel
    #print(part_number_ask)  # print part number in cmd to track parsing process
    # create full link to search for data
    url = 'https://exist.ua/price.aspx?pcode={}'.format(part_number_ask)
    nmb = randint(0, 3)  # create random index to set random header from headers
    # dictionary
    response = session.get(url, headers=headers[nmb])  # transfer server's
    # response to BS-object
    time.sleep(randint(3, 6))  # make artificial pause in algorithm to send
    # requests to site in moderate pace wihout overload
    if response.status_code != 200:  # check server's response - code 200
    # means everything is OK, otherwise print error
        print("{} not found (or page doesn't respond)".format(part_number_ask))
        # and store unfound requested part number to main dictionary with empty
        # N\A columns
        temporary.append({'stock': 'N/A', 'date': 'N/A', 'price': 0})
        data.append(
            {
                'part_number_ask': part_number_ask,
                'part_number_bid': 'N/A',
                'brand_bid': 'N/A',
                'description': 'N/A',
                'stock_date_price_div': temporary
            }
        )
    soup = BS(response.content, "html.parser")  # "response.content" - is what
    # must be transformed to BS-object, аnd "html.parser" - the way it must be
    # transformed
    div = soup.find('div', attrs={'id': 'priceBody'})  # manually find unique
    # <div> (by id "priceBody"), which stores needed data and indicate it for
    # the parser to cut searching area
    if div:
        # knowing that needed data can be stored on the first (preliminary) page
        # in <tr> tags, check what is the response from site - first or second
        # (final) page
        next_page = ''  # create empty variable to store address for final page
        trs = div.find_all('tr')
        for tr in trs:
            td = tr.find('td')  # brand name always is in first <td> so take it
            if brand_ask in td.a.text:  # check in <td> text "Mitsubishi"
                next_page = domain + td.a['href']  # if text "Mitsubishi" is in
                # <td>, take second part of the link and add it to domain to get
                # final link
                time.sleep(randint(3, 6)) # make artificial pause in algorithm
                # to send requests to site in moderate pace wihout overload
                nmb = randint(0, 3) # create random index to set random header
                # from headers dictionary
                # parse next (final) page
                response = session.get(next_page, headers = headers[nmb])
                # transfer server's response to BS-object
                soup = BS(response.content, "html.parser")
                # manually find unique <div> (by id "priceBody"), which stores
                # needed data and indicate it for parser to cut searching area
                div = soup.find('div',attrs={'id': 'priceBody'})

                parse_div(div)

        parse_div(div)

try:
    write_to_excel(data)
except Exception:
    print('There was a critical error while writing data to excel')
    logger.exception('There was a critical error while writing data to excel')
